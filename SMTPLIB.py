#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import os
import openpyxl
import smtplib
os.chdir(r"C:\Users\valikhan.baikanov\Desktop\Python\automate_online-materials")

# we use openpyxl module to loads excel source file
wb = openpyxl.load_workbook('example.xlsx') 
sheet = wb.get_sheet_by_name('Sheet1') # enables no operate Sheet1 via variable

# the code below is to setup the programm to use connect to email server and use email account
smtpObj = smtplib.SMTP('smtp.mail.ru', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('walikhan@mail.ru', 'ENTER_PASSWORD')

# the following loop goes through each row of column 5, finds empty cell, 
# and send email indicated in the same row and column 6
for i in range(1, 4):
	empt = sheet.cell(row=i, column=5).value
	if empt is None:
		smtpObj.sendmail('walikhan@mail.ru', sheet.cell(row=i, column=6).value, )
        'Subject: Delivery status.\nDear Supplier, please provide delivery status...')
	else smtpObj.quit()
    

